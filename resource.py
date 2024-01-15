#!/usr/bin/env python
# coding: utf-8

#-----------------------------------------------------------------------------------------------
# Script: Methods for metrics alert
# Description: Definations for methods provided in the configurations 
#
# Date: 12th May, 2023
# Initial Author: Yash Kolhatkar
# Domain Expert : 
# 
#
# Revisions:
# Date         |   Author              | Comments
#-----------------------------------------------------------------------------------------------


import pandas as pd
from datetime import datetime, timedelta
import logging
import os
import xlsxwriter
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',
                    filename='/mnt/c/Users/yash_kolhatkar/Desktop/Cannary/REPOSITORY/Logs/app.log')
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s %(message)s',
                    filename='/mnt/c/Users/yash_kolhatkar/Desktop/Cannary/REPOSITORY/Logs/app.log')

# Creating an object 
logger = logging.getLogger() 
# Setting the threshold of logger to DEBUG 
logger.setLevel(logging.DEBUG)


def calculate_diminishing_percentage(lst):
    if not lst:
        return "No metrics values present"

    num = lst[0]
    greater = 0
    smaller = 0
    equal = 0

    for i in lst:
        if num < i:
            num = i
            greater += 1
        elif num > i:
            num = i
            smaller += 1
        elif num == i:
            num = i
            equal += 1
    if greater > smaller:
        return "Increasing" 
    elif greater < smaller:
        return "Decreasing"
    elif equal > 0 and smaller == 0 and greater == 0:
        return f"No Change"
    else:
        return "Uneven"


def get_last_three_months(project_df, project_id, metrics_name):
    dt = datetime.now()
    logging.info("timestamp at the start of three months function: {}".format(dt))
    logging.info("inside get_last_three_months function")
    today = datetime.today().strftime('%b')
    last_three_months = [(datetime.today() - timedelta(days=30*i)).strftime('%b') for i in range(4)][::-1]
    logging.debug(f"last three months: {last_three_months}")
    
    # Create a pivoted dataframe where each row corresponds to a unique combination of Project ID, Metrics Name, and Month
    pivoted_df = project_df.pivot_table(index=['Project ID', 'Derived Metrics Name', 'Month'], values='Metrics Data').reset_index()
    logging.debug("Pivoted df created")

    # Filter the pivoted dataframe using the provided project_id and metrics_name
    filtered_df = pivoted_df[(pivoted_df['Project ID'] == project_id) & (pivoted_df['Derived Metrics Name'] == metrics_name)]
    logging.debug("filtered_df created")
    
    # Retrieve the metric values for the last three months
    metric_values = filtered_df[filtered_df['Month'].isin(last_three_months)].sort_values(by='Month', key=lambda x: pd.to_datetime(x, format='%b'))['Metrics Data'].tolist()
    logging.debug(f"returning metrics values for last three months: {metric_values}")
    dt = datetime.now()
    logging.info("timestamp at the end of three month function: {}".format(dt))
    return metric_values


def calculate_high_percentage(lst, min_thresh, max_thresh):
    # Check if the list is empty
    if not lst:
        return "No metrics values present for respective metrics"
    
    # Calculate the number of values in the list
    num_values = len(lst)
    
    # Count the number of values within the given threshold range
    num_values_in_range = sum(1 for val in lst if min_thresh <= val <= max_thresh)
    
    # Calculate the percentage of values within the given threshold range
    percentage_in_range = (num_values_in_range / num_values) * 100
    
    # Determine the alert message based on the percentage of values within the threshold range
    if percentage_in_range == 0:
        alert = "No values within range"
    elif percentage_in_range < 50:
        alert = f"{percentage_in_range}% of values within range (below threshold)"
    elif percentage_in_range == 100:
        alert = "All values within range"
    else:
        alert = f"{percentage_in_range}% of values within range (above threshold)"
    
    # Return the alert message
    return alert


def dummy_method():
    return f"dummy method for POC"


def replace_all(message, dic):
    for i, j in dic.items():
        message = message.replace(i, j)
    return message


def prepare_data(project_config_data, existing_data, column_df):
    
    new_rows = []
    for data in project_config_data:
       
        alert = data['config']['Engineering metrics trend']
        calculation_method = data['config']['Calculation Method']
        metrics_values = data['config']['Metrics Values']
        # values_no = len(metrics_values)
        calculation_method = calculation_method.replace("3", str(4))
        message_to_be_generated = data['config']['Message to be generated']
        metrics_values_string = ', '.join(str(value) for value in metrics_values)
        direction = data['config']['Direction']

        inplace_text = {
           "<Trend_Engg_Metrics> ": alert,
           "<Engineering Metrics>": data['Metrics Name'],
           "last 3 months": calculation_method,
           "M1,M2,M3": metrics_values_string
        }
        custom_message = replace_all(message_to_be_generated, inplace_text)
        trend_outcome = data['config']['Trend Outcome'] if data['config']['Trend Outcome'] else 'NaN'
        threshold_output = data['config']['Threshold Output']
        custom_message = custom_message + threshold_output

        if [data['Project ID'], data['Project Name'], data['Metrics Name'], data['config']['Derived Metrics Name'],
            metrics_values_string, alert, custom_message, trend_outcome, direction] in existing_data:
            continue
        
        row_data = {'Project ID': data['Project ID'],
                    'Project Name':data['Project Name'],
                    'Metrics Name':data['Metrics Name'],
                    'Derived Metrics Name':data['config']['Derived Metrics Name'],
                    'Metrics Values':metrics_values_string,
                    'Engineering metrics trend':alert,
                    'Custom message':custom_message,
                    'Trend Outcome':trend_outcome,
                    'Direction' : direction
                }
        new_rows.append(row_data)
        existing_data.append([
            data['Project ID'], data['Project Name'], data['Metrics Name'], data['config']['Derived Metrics Name'],
            metrics_values_string, alert, custom_message, trend_outcome, direction
            ])
    new_rows_df = pd.DataFrame(new_rows)
    project_data_df = pd.concat([column_df, new_rows_df], ignore_index=True)
    return project_data_df

    
def write_json_to_excel(project_config_data):
    column_names = ['Project ID', 'Project Name', 'Metrics Name', 'Derived Metrics Name', 'Metrics Values',
                   'Engineering metrics trend', 'Custom message', 'Trend Outcome', 'Direction']
 
    filename = '/mnt/c/Users/yash_kolhatkar/Desktop/Cannary/REPOSITORY/Data/target/new_metrics_analysis_report1.xlsx'
    sheet_name = "Engg Metric Analysis"

    workbook = Workbook()
    try:
        workbook = load_workbook(filename, read_only=False)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.workbook = workbook
        existing_data = workbook[sheet_name].values.tolist()
    except FileNotFoundError:
        workbook = Workbook()
        workbook.active.title = sheet_name
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.workbook = workbook
        bold_font = Font(bold=True)
        sheet = workbook[sheet_name]
        for col_num, column in enumerate(column_names, 1):
            cell = sheet.cell(row=1, column=col_num, value=column)
            cell.font = bold_font
        existing_data = []
    column_df = pd.DataFrame(columns = column_names)  
    df = prepare_data(project_config_data, existing_data, column_df)  # Assuming you have a function 'prepare_data' to get the data in DataFrame format
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

    return writer


def analyze_metrics_threshold(correlation, LSL, USL, metrics_values):
    # Check if LSL value is provided, if not, set to None
    LSL_present = False if str(LSL) == 'nan' else True

    # Check if USL value is provided, if not, set to None
    USL_present = False if str(USL) == 'nan' else True

    # Initialize variables to store the result messages
    metrics_value_msg = ""
    LSL_USL_msg = ""

    # Check if metrics value is within the range of USL and LSL
    if LSL_present and USL_present:
        if LSL <= metrics_values[-1] <= USL:
            metrics_value_msg = "Positive"
            LSL_USL_msg = " Metrics Value is {} within the range of {} to {}.".format(metrics_value_msg, LSL, USL)
        else:
            metrics_value_msg = "Negative"
            LSL_USL_msg = " Metrics Value is {} outside the range of {} to {}.".format(metrics_value_msg, LSL, USL)
    # Check Direct correlation conditions
    elif correlation == "Direct":
        if LSL_present:
            if metrics_values[-1] >= LSL:
                metrics_value_msg = "Positive"
                LSL_USL_msg = " Metrics Value is {}, It's above {}.".format(metrics_value_msg, LSL)
            else:
                metrics_value_msg = "Negative"
                LSL_USL_msg = " Metrics Value is {}, It's below {}.".format(metrics_value_msg, LSL)
        elif USL_present:
            if metrics_values[-1] >= USL:
                metrics_value_msg = "Positive"
                LSL_USL_msg = " Metrics Value is {}, It's above {}.".format(metrics_value_msg, USL)
            else:
                metrics_value_msg = "Negative"
                LSL_USL_msg = " Metrics Value is {}, It's below {}.".format(metrics_value_msg, USL)
        else:
            LSL_USL_msg = " No LSL and USL range available"
    # Check Indirect correlation conditions
    elif correlation == "Indirect":
        if LSL_present:
            if metrics_values[-1] < LSL:
                metrics_value_msg = "Positive"
                LSL_USL_msg = " Metrics Value is {}, It's below {}.".format(metrics_value_msg, LSL)
            else:
                metrics_value_msg = "Negative"
                LSL_USL_msg = " Metrics Value is {}, It's above {}.".format(metrics_value_msg, LSL)
        elif USL_present:
            if metrics_values[-1] > USL:
                metrics_value_msg = "Negative"
                LSL_USL_msg = " Metrics Value is {}, It's above {}.".format(metrics_value_msg, USL)
            else:
                metrics_value_msg = "Positive"
                LSL_USL_msg = " Metrics Value is {}, It's below {}.".format(metrics_value_msg, USL)
        else:
            LSL_USL_msg = " No LSL and USL range available"

    result_msg = LSL_USL_msg
    return result_msg
